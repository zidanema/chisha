"""随机抽 N 条 dishes_tagged 输出 markdown / xlsx 表格供人工抽查.

用法:
    uv run python -m scripts.sample_for_review home
    uv run python -m scripts.sample_for_review home --n 50 --seed 42
    uv run python -m scripts.sample_for_review home --format xlsx

默认 50 条, seed=42 (固定可复现, 抽查前后一致).
xlsx 格式需要 dev 依赖 openpyxl: `uv sync --extra dev`.
"""
from __future__ import annotations

import argparse
import json
import random
from pathlib import Path

from chisha.schemas import validate_dishes_tagged

ROOT = Path(__file__).resolve().parent.parent

# 抽查关键字段 (与 prompt 准则对齐, oil/veg/protein/complete/spicy 是 LLM 最易出错的)
# xlsx 输出比 md 多 cuisine / main_ingredient_type / cooking_method 3 列
HEADERS = [
    "dish_id",
    "raw_name",
    "canonical_name",
    "price",
    "cuisine",
    "main_ingredient",
    "cooking_method",
    "oil_level",
    "vegetable_ratio_estimate",
    "protein_grams_estimate",
    "is_complete_meal",
    "spicy_level",
]
# md 保留旧的 9 列 (向后兼容已有的 review 流程)
HEADERS_MD = [
    "dish_id",
    "raw_name",
    "canonical_name",
    "price",
    "oil_level",
    "vegetable_ratio_estimate",
    "protein_grams_estimate",
    "is_complete_meal",
    "spicy_level",
]


def _md_escape(s: str) -> str:
    """markdown table 里 | 必须 escape, 换行截断."""
    return str(s).replace("|", "\\|").replace("\n", " ").strip()


def render_row(d: dict) -> str:
    np = d["nutrition_profile"]
    cells = [
        d["dish_id"],
        d["raw_name"],
        d["canonical_name"],
        f"{d['price']:.1f}",
        np["oil_level"],
        f"{np['vegetable_ratio_estimate']:.2f}",
        np["protein_grams_estimate"],
        "✓" if np["is_complete_meal"] else "",
        np["spicy_level"],
    ]
    return "| " + " | ".join(_md_escape(c) for c in cells) + " |"


def render_markdown(sample: list[dict], zone: str, seed: int) -> str:
    lines = [
        f"# 打标抽查样本 — {zone}",
        "",
        f"- 样本量: {len(sample)} / {{total}}  (随机, seed={seed})",
        "- 来源: `data/{zone}/dishes_tagged.json`".replace("{zone}", zone),
        "- 校对方式: 逐条人眼判断关键字段 (oil_level / vegetable_ratio / "
        "protein_grams / is_complete_meal / spicy_level), 准确率 < 80% 回去改 prompt 重跑.",
        "",
        "| " + " | ".join(HEADERS_MD) + " |",
        "| " + " | ".join(["---"] * len(HEADERS_MD)) + " |",
    ]
    for d in sample:
        lines.append(render_row(d))
    lines.append("")
    return "\n".join(lines)


def _row_cells_full(d: dict) -> list:
    """xlsx 用: 全字段 + 末尾留 verdict / note 两列给人工填."""
    np = d["nutrition_profile"]
    return [
        d["dish_id"],
        d["raw_name"],
        d["canonical_name"],
        round(d["price"], 1),
        d.get("cuisine", ""),
        np["main_ingredient_type"],
        np["cooking_method"],
        np["oil_level"],
        round(np["vegetable_ratio_estimate"], 2),
        np["protein_grams_estimate"],
        "✓" if np["is_complete_meal"] else "",
        np["spicy_level"],
    ]


def write_xlsx(sample: list[dict], zone: str, seed: int, total: int,
               out_path: Path) -> None:
    """写 xlsx: 冻结首行 + 自适应列宽 + auto filter + 末尾加 verdict/note 列."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit(
            "需要 openpyxl, 执行: uv sync --extra dev"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"review_{zone[:20]}"

    # 标题行 = HEADERS + 人工填写的两列
    headers = HEADERS + ["verdict", "note"]
    ws.append(headers)

    # 表头加粗 + 浅灰底
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E0E0E0")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for d in sample:
        ws.append(_row_cells_full(d) + ["", ""])

    # 列宽自适应 (按内容长度 +2, 限制 50)
    for col_idx, name in enumerate(headers, start=1):
        max_len = len(str(name))
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            v = row[0].value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            50, max_len + 2
        )

    # 冻结首行 + auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # 元信息写到第二个 sheet
    meta = wb.create_sheet("meta")
    meta.append(["zone", zone])
    meta.append(["sample_n", len(sample)])
    meta.append(["total_in_pool", total])
    meta.append(["seed", seed])
    meta.append(["source", f"data/{zone}/dishes_tagged.json"])
    meta.append(["verdict 用法", "OK / oil / veg / prot / meal / spicy / "
                 "cuisine / ingredient / cooking / name / multi"])
    meta.append(["验收标准", "准确率 < 80% 回去改 prompts/tag_dishes.md 重跑"])
    for col_idx in (1, 2):
        meta.column_dimensions[get_column_letter(col_idx)].width = 40

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("zone", help="office_zone, e.g. home")
    ap.add_argument("--n", type=int, default=50, help="抽样条数 (默认 50)")
    ap.add_argument("--seed", type=int, default=42, help="随机种子 (默认 42)")
    ap.add_argument("--format", choices=("md", "xlsx", "both"),
                    default="md", help="输出格式 (默认 md)")
    ap.add_argument("--out", type=str, default=None,
                    help="输出路径 (默认 data/{zone}/review_sample.{ext})")
    args = ap.parse_args()

    base = ROOT / "data" / args.zone
    src = base / "dishes_tagged.json"
    if not src.exists():
        raise SystemExit(f"missing: {src}")

    records = json.loads(src.read_text(encoding="utf-8"))
    # 跑一遍 schema, 抽查样本必然合法
    validate_dishes_tagged(records)

    if len(records) < args.n:
        raise SystemExit(
            f"records ({len(records)}) < n ({args.n}); 调小 --n 或先跑全量打标"
        )

    rng = random.Random(args.seed)
    sample = rng.sample(records, args.n)

    formats = ("md", "xlsx") if args.format == "both" else (args.format,)
    for fmt in formats:
        if args.out and len(formats) == 1:
            out_path = Path(args.out)
        else:
            out_path = base / f"review_sample.{fmt}"
        if fmt == "md":
            md = render_markdown(sample, args.zone, args.seed).replace(
                "{total}", str(len(records))
            )
            out_path.write_text(md, encoding="utf-8")
        else:
            write_xlsx(sample, args.zone, args.seed, len(records), out_path)
        print(f"wrote {args.n}-row review sample → {out_path}")


if __name__ == "__main__":
    main()
